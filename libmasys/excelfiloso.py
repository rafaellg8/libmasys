#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE','libmasys.settings')

import django
from openpyxl import load_workbook

django.setup()

from bookstore.models import Genero,Recurso,Usuario,Prestamo
"""
FILOSOFIA
"""
genero = Genero(nombre="Filosofía y pensamiento")
genero.save()

wb2 = load_workbook('Filosofia.xlsx')
ws = wb2.active


for row in ws['A2':'H183']:
        recurso = Recurso(titulo=row[0].value,autor=row[1].value,anio=str(int(row[3].value)),editorial=row[4].value,genero=genero,codigo=row[7].value)
        recurso.save()

for row in ws['A199':'H247']:
        recurso = Recurso(titulo=row[0].value,autor=row[1].value,anio=str(int(row[3].value)),editorial=row[4].value,genero=genero,codigo=row[7].value)
        recurso.save()
