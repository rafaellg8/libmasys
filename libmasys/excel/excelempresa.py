#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE','libmasys.settings')

import django
from openpyxl import load_workbook

django.setup()

from bookstore.models import Genero,Recurso,Usuario,Prestamo
"""
FILOSOFIA
"""
genero = Genero(nombre="Economía - Ciencias Empresariales")
genero.save()

recursos = Recurso.objects.filter(genero=genero)
if len(recursos)>0:
    for r in recursos:
        r.delete()


wb2 = load_workbook('Empresariales.xlsx')
ws = wb2.active


for row in ws['A63':'H125']:
        recurso = Recurso(titulo=row[0].value,autor=row[1].value,anio=str(int(row[3].value)),editorial=row[4].value,genero=genero,codigo=row[7].value)
        recurso.save()

# for row in ws['A252':'H293']:
#         recurso = Recurso..objects.update_or_create(titulo=row[0].value,autor=row[1].value,anio=str(int(row[3].value)),editorial=row[4].value,genero=genero,codigo=row[7].value)
#         recurso.save()
